# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
AutoNuclei 自动化扫描工具

该工具用于自动化执行资产扫描和漏洞检测流程，主要功能包括：
1. 调用EHole进行指纹识别
2. 解析Excel格式的扫描结果
3. 根据指纹识别结果，使用Nuclei进行漏洞扫描
4. 支持基于风险级别和标签的扫描策略

作者: base64_painter
团队: base64-sec
博客: https://painter-sec.cnblogs.com
"""
from openpyxl import workbook, load_workbook
import os
import platform
import time
import json


import ProcessMonitor

class AutoNuclei():
    """
    AutoNuclei类用于自动化执行资产扫描和漏洞检测流程
    该类封装了从指纹识别到漏洞扫描的完整流程
    """

    def __init__(self):
        """
        初始化AutoNuclei实例并启动完整的扫描流程
        流程包括：显示标题、创建目录、启动指纹扫描、加载配置、解析数据、执行扫描
        """
        self.plat = platform.system().lower()
        if self.plat == 'windows':
            self.nuclei_path = '.\\module\\Nuclei\\nuclei.exe'
            self.tags_file = '.\\nuclei-templates\\TEMPLATES-STATS.json'  # 读取nuclei里所有的tags标签
        else:
            self.nuclei_path = os.environ['HOME'] + '/nuclei'
            self.tags_file = './nuclei-templates/TEMPLATES-STATS.json'  # 读取nuclei里所有的tags标签
        self.target_file = 'wait_check.xlsx'  # 默认，不要动

        self.level_target_temp = 'temp/level_target_temp.txt'  # 存储用于等级扫描的目标临时文件
        self.tag_target_temp = 'temp/tag_target_temp.txt'  # 存储用于标签扫描的目标临时文件
        # self.nuclei_path = r'C:\Users\painter\Desktop\auto_nuclei'
        self.nuclei_tags = []  # 存放nuclei里的所有tags
        self.all_message = []  # 存放从Excel读取的所有数据
        self.base_scan_target = []  # 存放后续用高低中等级的poc来进行扫描的目标
        self.tags_scan_target = []  # 存放可以用nuclei的tags来进行扫描的目标

        self.create_dir()  # 创建必要的目录
        self.start_Ehole()  # 启动资产扫描(EHole指纹识别)
        self.get_nuclei_tags()  # 获取nuclei的所有标签
        self.excel_load()  # 加载Excel文件数据
        self.extract_data()  # 根据excel资产，来采用不同的方式调用nuclei

        self.tag_scan()  # 启用标签扫描
        self.level_scan()  # 启用高低中扫描(当前注释掉)
        # 启动 xray + rad
        self.xray_rad()

    def create_dir(self):
        """
        创建临时目录和结果目录
        确保temp和result目录存在，用于存储临时文件和扫描结果
        """
        if not os.path.exists('temp'):
            os.mkdir('temp')
        if not os.path.exists('result'):
            os.mkdir('result')
        time.sleep(1)  # 等待目录创建完成

    def start_Ehole(self):
        """
        启动EHole工具进行指纹识别
        调用EHole.exe对targets.txt中的目标进行指纹识别，并将结果保存到wait_check.xlsx
        使用ProcessMonitor监控EHole进程，等待其执行完成
        """
        cmd = 'module\\EHole\\EHole.exe finger -l targets.txt -o wait_check.xlsx'
        print('[+]正在进行指纹扫描...')
        os.system(cmd)
        while True:
            status = ProcessMonitor.ProcessMonitor("EHole").main()  # 获取EHole是否执行完毕
            if status:  # 如果执行完毕，则跳出循环
                break

    def get_nuclei_tags(self):
        """
        从nuclei模板统计文件中获取所有可用的标签
        读取TEMPLATES-STATS.json文件，提取其中的所有标签名称并存储到self.nuclei_tags列表中
        """
        temp = json.load(open(self.tags_file))
        messages = temp["tags"]
        for info in messages:
            if info['name'] != '':  # 过滤掉空标签
                self.nuclei_tags.append(info['name'])  # 将所有的标签储存起来

    def excel_load(self):
        """
        加载并读取Excel格式的扫描结果文件
        打开wait_check.xlsx文件，读取Sheet1中的所有数据行并存储到self.all_message列表中
        """
        # 读取工作簿
        wb = load_workbook(self.target_file)  # 打开工作簿
        # 获取sheet，获取所有sheet
        # 获取第一张表
        sheet = wb['Sheet1']  # 这是我们的需要的表
        for row in sheet.iter_rows(min_row=2):  # 遍历,从第2行开始读取所有行
            row_message = []  # 存放每一行的数据
            for cel in row:
                if row[0].value == None:  # 没数据了，就退出
                    break
                row_message.append(cel.value)  # 每一行的数据
            self.all_message.append(row_message)

    # 对资产进行分类
    def extract_data(self):
        """
        根据Excel中的资产数据进行分类处理
        将目标分为两类：
        1. 有匹配指纹标签的目标，存储到self.tags_scan_target字典(标签->目标列表)
        2. 无匹配指纹标签的目标，存储到self.base_scan_target列表，后续使用高低中等级POC扫描
        """
        # print(self.all_message)
        temp_result = []
        for message in self.all_message:
            url = {}  # 存储当前目标匹配的标签和URL
            stat = 0  # 开关，表示是否匹配到标签

            if message[1] == '':  # 如果没有指纹信息
                self.base_scan_target.append(message[0])  # 该目标后续用高低中poc来进行批量扫描
            else:
                # 检查目标指纹是否匹配任何nuclei标签
                for t in self.nuclei_tags:
                    if str(t).lower() in str(message[1]).lower():
                        stat = 1
                        url[t] = message[0]
                if stat == 0:  # 一个目标，一个标签都没命中，就让其进行高低中扫描
                    self.base_scan_target.append(message[0])  # 该目标后续用高低中poc来进行批量扫描
            temp_result.append(url)

        # 整合结果，将相同标签的目标合并到一起
        result = {}
        for temp in temp_result:
            for tag in temp.keys():
                if result.get(tag) == None:
                    result.update({tag: []})
                for t in temp.values():
                    result.get(tag).append(t)
        self.tags_scan_target = result

    def level_scan(self):
        """
        执行基于风险级别的扫描
        将self.base_scan_target中的目标写入临时文件，然后使用nuclei进行high和critical级别的漏洞扫描
        扫描结果保存到result/level_vul_result.txt文件
        """
        with open(self.level_target_temp, 'a+', encoding="utf-8") as f:
            # print(self.base_scan_target)
            for info in self.base_scan_target:
                f.write(info + '\n')
        # 构建nuclei命令，使用high和critical级别的模板进行扫描
        cmd = (f'{self.nuclei_path}'
               f' -l {self.level_target_temp}'
               f' -s high,critical'
               f' -o result/level_vul_result.txt'
               f' -stats'
               f' -t ./nuclei-templates'
               f' -duc')

        # 等tag_scan 结束
        status = ProcessMonitor.ProcessMonitor("nuclei").main()
        print(cmd)
        if status:
            os.system(cmd)

    def tag_scan(self):
        """
        执行基于标签的扫描
        对self.tags_scan_target中的每个标签及其对应的目标列表，分别执行nuclei扫描
        每个标签的扫描结果保存到单独的文件中
        """
        # 遍历每个标签及其对应的目标列表
        for tag_name, targets in self.tags_scan_target.items():  # 解包，获取标签名，和每个标签对应的所有资产
            with open(self.tag_target_temp, 'a+', encoding="utf-8") as f:
                for target in targets:
                    f.write(target + '\n')
            # 构建nuclei命令，使用特定标签的模板进行扫描
            cmd = (f'{self.nuclei_path}'
                   f' -l {self.tag_target_temp}'
                   f' -tags {tag_name}'
                   f' -stats '
                   f' -o result/tag_vul_{time.time()}.txt'
                   f' -t ./nuclei-templates'
                   f' -duc'
                   )
            print(cmd)
            os.system(cmd)

    def xray_rad(self):
        return None

if __name__ == '__main__':
    """
    程序入口点
    创建AutoNuclei实例并启动扫描流程
    """
    scan = AutoNuclei()
